"""
Text extraction for CSV and XLSX files.
CSV: converts rows to readable text.
XLSX: reads all sheets and converts to text with sheet name headers.
"""

import csv
import io
from openpyxl import load_workbook


def extract_text_from_csv(file_path: str) -> str:
    """Extract text from a CSV file, converting rows to readable lines."""
    rows = []
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return ""

        rows.append(" | ".join(header))
        rows.append("-" * len(rows[0]))

        for row in reader:
            if any(cell.strip() for cell in row):
                rows.append(" | ".join(row))

    return "\n".join(rows)


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from an XLSX file, reading all sheets."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))

        if rows:
            header = f"--- Sheet: {sheet_name} ---"
            sections.append(f"{header}\n{chr(10).join(rows)}")

    wb.close()
    return "\n\n".join(sections)
