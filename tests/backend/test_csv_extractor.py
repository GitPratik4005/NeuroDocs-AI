"""Unit tests for CSV and XLSX text extraction."""

import os
import csv
import pytest
from openpyxl import Workbook
from services.csv_extractor import extract_text_from_csv, extract_text_from_xlsx


@pytest.fixture
def tmp_csv(tmp_path):
    """Create a temporary CSV file."""
    path = str(tmp_path / "test.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Age", "City"])
        writer.writerow(["Alice", "30", "New York"])
        writer.writerow(["Bob", "25", "London"])
    return path


@pytest.fixture
def tmp_csv_empty(tmp_path):
    """Create an empty CSV file."""
    path = str(tmp_path / "empty.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        pass
    return path


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a temporary XLSX file with two sheets."""
    path = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Product", "Revenue"])
    ws1.append(["Widget", "1000"])
    ws1.append(["Gadget", "2000"])

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", "500"])

    wb.save(path)
    return path


@pytest.fixture
def tmp_xlsx_empty(tmp_path):
    """Create an XLSX file with an empty sheet."""
    path = str(tmp_path / "empty.xlsx")
    wb = Workbook()
    wb.save(path)
    return path


def test_csv_extraction(tmp_csv):
    text = extract_text_from_csv(tmp_csv)
    assert "Name" in text
    assert "Alice" in text
    assert "Bob" in text
    assert "New York" in text


def test_csv_has_header_separator(tmp_csv):
    text = extract_text_from_csv(tmp_csv)
    lines = text.split("\n")
    assert len(lines) >= 3
    assert "---" in lines[1]


def test_csv_empty_file(tmp_csv_empty):
    text = extract_text_from_csv(tmp_csv_empty)
    assert text == ""


def test_xlsx_extraction(tmp_xlsx):
    text = extract_text_from_xlsx(tmp_xlsx)
    assert "Sales" in text
    assert "Widget" in text
    assert "1000" in text
    assert "Expenses" in text
    assert "Rent" in text


def test_xlsx_multi_sheet(tmp_xlsx):
    text = extract_text_from_xlsx(tmp_xlsx)
    assert "--- Sheet: Sales ---" in text
    assert "--- Sheet: Expenses ---" in text


def test_xlsx_empty_sheet(tmp_xlsx_empty):
    text = extract_text_from_xlsx(tmp_xlsx_empty)
    # Empty sheet should produce no content
    assert text == ""
